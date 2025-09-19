import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading
import csv
import time
import re
import os 
import json
import webbrowser
import traceback
import random
from datetime import datetime
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError, Error as PlaywrightError

# Import for Gemini AI
try:
    import google.generativeai as genai
    HAS_GEMINI = True
except ImportError:
    HAS_GEMINI = False

# Try to import optional libraries
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

class MultiStoreScraperGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Multi-Store Price Scraper")
        self.root.geometry("1200x700")

        self.scraped_data = []
        self.is_scraping = False
        self.api_key = tk.StringVar()
        self.model_var = tk.StringVar()
        self.available_models = ['gemini-2.5-flash', 'gemini-2.5-pro']
        
        # File for storing URLs
        self.urls_file = "scraper_urls.txt"

        self.default_urls = [
            # Woolworths URLs
            'https://www.woolworths.com.au/shop/productdetails/160209/restor-concentrated-laundry-detergent-sheets-fresh-linen',
            'https://www.woolworths.com.au/shop/productdetails/164887/restor-concentrated-laundry-detergent-sheets-tropical',
            'https://www.woolworths.com.au/shop/productdetails/899864/undo-this-mess-laundry-detergent-sheets-spring-blossom',
            'https://www.woolworths.com.au/shop/productdetails/272748/earth-rescue-laundry-detergent-sheets-60-loads',
            'https://www.woolworths.com.au/shop/productdetails/909874/sheet-yeah-laundry-detergent-sheets-summer-daze',
            'https://www.woolworths.com.au/shop/productdetails/897214/dr-beckmann-laundry-detergent-sheets-universal',
            'https://www.woolworths.com.au/shop/productdetails/1122347429/cleaner-days-laundry-detergent-sheets-fragrance-free-60-pack',
            'https://www.woolworths.com.au/shop/productdetails/1122351339/cleaner-days-laundry-detergent-sheets-lemon-eucalyptus-60-pack',
            'https://www.woolworths.com.au/shop/productdetails/6021253/restor-dishwasher-detergent-sheets-lemon-burst',
            'https://www.woolworths.com.au/shop/productdetails/6020374/lucent-globe-dishwashing-detergent-sheets-fresh-lemon-scent',
            'https://www.woolworths.com.au/shop/productdetails/677834/earth-rescue-dishwasher-detergent-sheets-lemon-large',
            'https://www.woolworths.com.au/shop/productdetails/685913/earth-rescue-dishwasher-detergent-sheets-lemon-large',
            'https://www.woolworths.com.au/shop/productdetails/1122351535/cleaner-days-dishwasher-detergent-sheets-lemon-60-pack',
            # Coles URLs
            'https://www.coles.com.au/product/undo-this-mess-laundry-detergent-sheets-spring-blossom-scent-60-pack-8415795',
            'https://www.coles.com.au/product/dr.-beckmann-magic-leaves-laundry-detergent-sheets-universal-25-pack-5452994',
            'https://www.coles.com.au/product/ecostore-laundry-detergent-sheets-fragrance-free-40-pack-8924623',
            'https://www.coles.com.au/product/lucent-globe-dishwashing-detergent-sheets-fresh-lemon-35-pack-1066354'
        ]

        self.setup_ui()
        self.load_settings()
        self.load_urls_from_file()

    def setup_ui(self):
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill='both', expand=True, padx=5, pady=5)
        self.urls_frame = ttk.Frame(notebook)
        notebook.add(self.urls_frame, text='URLs')
        self.setup_urls_tab()
        self.results_frame = ttk.Frame(notebook)
        notebook.add(self.results_frame, text='Results')
        self.setup_results_tab()
        self.log_frame = ttk.Frame(notebook)
        notebook.add(self.log_frame, text='Log')
        self.setup_log_tab()
        self.settings_frame = ttk.Frame(notebook)
        notebook.add(self.settings_frame, text='Settings')
        self.setup_settings_tab()
        self.setup_control_panel()

    def setup_urls_tab(self):
        list_frame = ttk.Frame(self.urls_frame)
        list_frame.pack(fill='both', expand=True, padx=10, pady=10)
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side='right', fill='y')
        self.url_listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set, height=15)
        self.url_listbox.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=self.url_listbox.yview)
        
        instructions = ttk.Label(self.urls_frame, text="Add Woolworths or Coles product URLs. The tool will automatically detect the store.\nURLs are automatically saved to 'scraper_urls.txt'", wraplength=1000)
        instructions.pack(pady=5, padx=10)
        
        controls_frame = ttk.Frame(self.urls_frame)
        controls_frame.pack(fill='x', padx=10, pady=5)
        entry_frame = ttk.Frame(controls_frame)
        entry_frame.pack(fill='x', pady=(0, 5))
        self.url_entry = ttk.Entry(entry_frame)
        self.url_entry.pack(side='left', expand=True, fill='x', padx=(0, 5))
        ttk.Button(entry_frame, text="Add URL", command=self.add_url).pack(side='left')
        
        button_frame = ttk.Frame(controls_frame)
        button_frame.pack(fill='x', expand=True)
        button_frame.columnconfigure((0, 1, 2, 3, 4), weight=1)
        ttk.Button(button_frame, text="Remove Selected", command=self.remove_url).grid(row=0, column=0, sticky='ew', padx=2)
        ttk.Button(button_frame, text="Clear All", command=self.clear_urls).grid(row=0, column=1, sticky='ew', padx=2)
        ttk.Button(button_frame, text="Load Defaults", command=self.reset_urls).grid(row=0, column=2, sticky='ew', padx=2)
        ttk.Button(button_frame, text="Import URLs", command=self.import_urls).grid(row=0, column=3, sticky='ew', padx=2)
        ttk.Button(button_frame, text="Export URLs", command=self.export_urls).grid(row=0, column=4, sticky='ew', padx=2)

    def setup_results_tab(self):
        columns = ('Store', 'Product', 'Price', 'Was', 'Unit Price', 'Promotion')
        self.tree = ttk.Treeview(self.results_frame, columns=columns, show='headings', height=20)
        self.tree.heading('Store', text='Store')
        self.tree.heading('Product', text='Product Name')
        self.tree.heading('Price', text='Current Price')
        self.tree.heading('Was', text='Was Price')
        self.tree.heading('Unit Price', text='Unit Price')
        self.tree.heading('Promotion', text='Promotion')
        self.tree.column('Store', width=80, anchor='center')
        self.tree.column('Product', width=400)
        self.tree.column('Price', width=100, anchor='center')
        self.tree.column('Was', width=100, anchor='center')
        self.tree.column('Unit Price', width=150, anchor='center')
        self.tree.column('Promotion', width=150, anchor='center')
        self.tree.bind("<Double-1>", self.on_item_double_click)
        vsb = ttk.Scrollbar(self.results_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(self.results_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(column=0, row=0, sticky='nsew', padx=5, pady=5)
        vsb.grid(column=1, row=0, sticky='ns')
        hsb.grid(column=0, row=1, sticky='ew')
        self.results_frame.grid_columnconfigure(0, weight=1)
        self.results_frame.grid_rowconfigure(0, weight=1)
        self.summary_label = ttk.Label(self.results_frame, text="No data scraped yet")
        self.summary_label.grid(column=0, row=2, pady=5, sticky='w', padx=5)

    def setup_settings_tab(self):
        settings_pane = ttk.Frame(self.settings_frame, padding="10")
        settings_pane.pack(fill='x', expand=False, padx=10, pady=10)
        ttk.Label(settings_pane, text="Gemini API Key:", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky='w', pady=5)
        api_key_instructions = "Get your API key from Google AI Studio. Settings are saved locally in a config.json file."
        ttk.Label(settings_pane, text=api_key_instructions, wraplength=500).grid(row=1, column=0, columnspan=2, sticky='w', pady=(0, 15))
        self.api_key_entry = ttk.Entry(settings_pane, textvariable=self.api_key, width=70, show='*')
        self.api_key_entry.grid(row=2, column=0, columnspan=2, sticky='ew', padx=(0, 5))
        ttk.Label(settings_pane, text="Gemini Model:", font=('Arial', 10, 'bold')).grid(row=3, column=0, sticky='w', pady=(20, 5))
        model_instructions = "Select the AI model to use. 'Flash' is faster and cheaper, while 'Pro' is more powerful for complex analysis."
        ttk.Label(settings_pane, text=model_instructions, wraplength=500).grid(row=4, column=0, columnspan=2, sticky='w', pady=(0, 10))
        self.model_selector = ttk.Combobox(settings_pane, textvariable=self.model_var, values=self.available_models, state="readonly")
        self.model_selector.grid(row=5, column=0, columnspan=2, sticky='ew')
        self.model_selector.set(self.available_models[0])
        self.save_settings_button = ttk.Button(settings_pane, text="Save Settings", command=self.save_settings)
        self.save_settings_button.grid(row=6, column=0, columnspan=2, pady=(20, 0))
        settings_pane.grid_columnconfigure(0, weight=1)

    def setup_log_tab(self):
        self.log_text = scrolledtext.ScrolledText(self.log_frame, height=25, width=100, wrap=tk.WORD)
        self.log_text.pack(fill='both', expand=True, padx=10, pady=10)

    def setup_control_panel(self):
        control_frame = ttk.Frame(self.root)
        control_frame.pack(fill='x', side='bottom', padx=10, pady=10)
        self.progress = ttk.Progressbar(control_frame, mode='determinate')
        self.progress.pack(fill='x', padx=5, pady=(0, 10))
        left_frame = ttk.Frame(control_frame)
        left_frame.pack(side='left', padx=5)
        store_frame = ttk.LabelFrame(left_frame, text="Stores")
        store_frame.pack(side='left', padx=(0, 10))
        self.scrape_woolworths = tk.BooleanVar(value=True)
        self.scrape_coles = tk.BooleanVar(value=True)
        ttk.Checkbutton(store_frame, text="Woolworths", variable=self.scrape_woolworths).pack(anchor='w')
        ttk.Checkbutton(store_frame, text="Coles", variable=self.scrape_coles).pack(anchor='w')
        options_frame = ttk.LabelFrame(left_frame, text="Options")
        options_frame.pack(side='left')
        self.headless_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(options_frame, text="Headless Mode", variable=self.headless_var).pack(anchor='w')
        self.debug_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(options_frame, text="Debug Mode", variable=self.debug_var).pack(anchor='w')
        buttons_frame = ttk.Frame(control_frame)
        buttons_frame.pack(side='right', padx=5)
        self.scrape_button = ttk.Button(buttons_frame, text="Start Scraping", command=self.start_scraping)
        self.scrape_button.pack(side='left', padx=5)
        self.csv_button = ttk.Button(buttons_frame, text="Export CSV", command=self.export_csv, state='disabled')
        self.csv_button.pack(side='left', padx=5)
        self.excel_button = ttk.Button(buttons_frame, text="Export Excel", command=self.export_excel, state='disabled')
        self.excel_button.pack(side='left', padx=5)
        self.ai_button = ttk.Button(buttons_frame, text="AI Analyse", command=self.start_ai_analysis, state='disabled')
        self.ai_button.pack(side='left', padx=5)

    def save_urls_to_file(self):
        """Save all URLs from the listbox to a text file"""
        try:
            urls = list(self.url_listbox.get(0, tk.END))
            with open(self.urls_file, 'w', encoding='utf-8') as f:
                for url in urls:
                    f.write(url + '\n')
            self.log(f"URLs saved to {self.urls_file}")
        except Exception as e:
            self.log(f"Error saving URLs to file: {e}")

    def load_urls_from_file(self):
        """Load URLs from the text file if it exists"""
        try:
            if os.path.exists(self.urls_file):
                with open(self.urls_file, 'r', encoding='utf-8') as f:
                    urls = [line.strip() for line in f if line.strip()]
                
                if urls:
                    self.url_listbox.delete(0, tk.END)
                    for url in urls:
                        if 'woolworths.com.au/shop/productdetails/' in url or 'coles.com.au/product/' in url:
                            self.url_listbox.insert(tk.END, url)
                    self.log(f"Loaded {len(urls)} URLs from {self.urls_file}")
                else:
                    # If file is empty, load defaults
                    self.load_default_urls()
            else:
                # If file doesn't exist, create it with defaults
                self.load_default_urls()
                self.save_urls_to_file()
        except Exception as e:
            self.log(f"Error loading URLs from file: {e}")
            self.load_default_urls()

    def load_default_urls(self):
        """Load the default URLs into the listbox"""
        for url in self.default_urls:
            self.url_listbox.insert(tk.END, url)

    def import_urls(self):
        """Import URLs from a user-selected text file"""
        filename = filedialog.askopenfilename(
            title="Import URLs from text file",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        
        if filename:
            try:
                with open(filename, 'r', encoding='utf-8') as f:
                    imported_urls = [line.strip() for line in f if line.strip()]
                
                imported_count = 0
                for url in imported_urls:
                    if 'woolworths.com.au/shop/productdetails/' in url or 'coles.com.au/product/' in url:
                        self.url_listbox.insert(tk.END, url)
                        imported_count += 1
                
                if imported_count > 0:
                    self.save_urls_to_file()
                    self.log(f"Imported {imported_count} valid URLs from {os.path.basename(filename)}")
                    messagebox.showinfo("Import Complete", f"Successfully imported {imported_count} URLs")
                else:
                    messagebox.showwarning("No Valid URLs", "No valid Woolworths or Coles URLs found in the file")
                    
            except Exception as e:
                messagebox.showerror("Import Error", f"Failed to import URLs: {e}")

    def export_urls(self):
        """Export current URLs to a user-selected text file"""
        urls = list(self.url_listbox.get(0, tk.END))
        if not urls:
            messagebox.showwarning("No URLs", "No URLs to export")
            return
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
            initialfile=f"exported_urls_{timestamp}.txt"
        )
        
        if filename:
            try:
                with open(filename, 'w', encoding='utf-8') as f:
                    for url in urls:
                        f.write(url + '\n')
                self.log(f"Exported {len(urls)} URLs to {os.path.basename(filename)}")
                messagebox.showinfo("Export Complete", f"Successfully exported {len(urls)} URLs")
            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to export URLs: {e}")

    def add_url(self):
        url = self.url_entry.get().strip()
        if url and ('woolworths.com.au/shop/productdetails/' in url or 'coles.com.au/product/' in url):
            self.url_listbox.insert(tk.END, url)
            self.url_entry.delete(0, tk.END)
            self.save_urls_to_file()  # Auto-save after adding
            self.log("Added URL: " + url)
        else:
            messagebox.showwarning("Invalid URL", "Please enter a valid Woolworths or Coles product URL.")

    def remove_url(self):
        selection = self.url_listbox.curselection()
        if selection:
            self.url_listbox.delete(selection)
            self.save_urls_to_file()  # Auto-save after removing
            self.log("Removed selected URL")

    def clear_urls(self):
        if messagebox.askyesno("Confirm Clear", "Are you sure you want to clear all URLs?"):
            self.url_listbox.delete(0, tk.END)
            self.save_urls_to_file()  # Auto-save after clearing
            self.log("Cleared all URLs")

    def reset_urls(self):
        if messagebox.askyesno("Load Defaults", "This will replace all current URLs with the default list. Continue?"):
            self.clear_urls()
            for url in self.default_urls:
                self.url_listbox.insert(tk.END, url)
            self.save_urls_to_file()  # Auto-save after resetting
            self.log("Loaded default URLs")

    def log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def calculate_discount(self, current_price, was_price, promo_badge=""):
        if was_price and was_price not in ["Not applicable", "-", ""]:
            try:
                current_match = re.search(r'[\d.]+', str(current_price))
                was_match = re.search(r'[\d.]+', str(was_price))
                if current_match and was_match:
                    current, was = float(current_match.group()), float(was_match.group())
                    if was > current:
                        discount_amt, discount_pct = was - current, ((was - current) / was) * 100
                        if "1/2 Price" in promo_badge or 48 <= discount_pct <= 52: return discount_pct, "HALF PRICE!"
                        if "Special" in promo_badge: return discount_pct, f"{discount_pct:.0f}% OFF"
                        if discount_pct >= 30: return discount_pct, f"{discount_pct:.0f}% OFF!"
                        if discount_pct >= 20: return discount_pct, f"{discount_pct:.0f}% OFF"
                        if discount_pct > 0: return discount_pct, f"Save ${discount_amt:.2f}"
            except Exception: pass
        if promo_badge:
            if "1/2 Price" in promo_badge: return 50.0, "HALF PRICE!"
            if "Special" in promo_badge: return None, "SPECIAL"
        return None, ""

    def _save_debug_html(self, page, store, url):
        try:
            html_content = page.content()
            sanitized_name = re.sub(r'[^a-zA-Z0-9_-]', '', url.split('/')[-1])[:50]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"debug_{store}_{sanitized_name}_{timestamp}.html"
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(html_content)
            self.log(f"  [DEBUG] Saved page HTML to {filename}")
        except Exception as e:
            self.log(f"  [DEBUG] Could not save HTML file: {e}")

    def save_cookies(self, context):
        """Save browser cookies for reuse"""
        try:
            cookies = context.cookies()
            with open('cookies.json', 'w') as f:
                json.dump(cookies, f)
            self.log("Cookies saved for next session")
        except Exception as e:
            self.log(f"Could not save cookies: {e}")

    def load_cookies(self, context):
        """Load saved cookies if available"""
        try:
            if os.path.exists('cookies.json'):
                with open('cookies.json', 'r') as f:
                    cookies = json.load(f)
                    context.add_cookies(cookies)
                self.log("Previous session cookies loaded")
                return True
        except Exception as e:
            self.log(f"Could not load cookies: {e}")
        return False

    def warmup_browser(self, page):
        """Simulate normal browsing behavior before scraping"""
        try:
            # Visit Coles homepage first
            page.goto('https://www.coles.com.au', wait_until='networkidle', timeout=30000)
            time.sleep(random.uniform(2, 4))
            
            # Scroll a bit to simulate human behavior
            page.evaluate("window.scrollTo(0, 300)")
            time.sleep(random.uniform(1, 2))
            
            page.evaluate("window.scrollTo(0, 0)")
            time.sleep(random.uniform(0.5, 1))
            
            self.log("Browser warmup completed")
        except Exception as e:
            self.log(f"Warmup failed (non-critical): {e}")

    def scrape_woolworths_page(self, page, url):
        page.goto(url, wait_until='domcontentloaded', timeout=30000)
        if self.debug_var.get():
            self._save_debug_html(page, 'woolworths', url)
        panel = page.locator('section[class*="product-details-panel_component_product-panel"]')
        panel.wait_for(timeout=20000)
        name = panel.locator('h1[class*="product-title_component_product-title"]').inner_text()
        price, was_price, cup_price, promo_badge = "Not found", "Not applicable", "Not found", ""
        try: price = panel.locator('div[class*="product-price_component_price-lead"]').inner_text(timeout=5000).replace('$', '').strip()
        except Exception: pass
        try: 
            was_price_full = panel.locator('div[class*="product-unit-price_component_price-was"]').inner_text(timeout=2000)
            was_price = re.search(r'[\d.]+', was_price_full).group()
        except Exception: pass
        try: cup_price = panel.locator('div[class*="product-unit-price_component_price-cup-string"]').inner_text(timeout=5000)
        except Exception: pass
        try: promo_badge = panel.locator('div[class*="product-stamp_message"]').inner_text(timeout=1000)
        except Exception: pass
        return {'store': 'Woolworths', 'name': name, 'price': price, 'was_price': was_price, 'cup_price': cup_price, 'url': url, 'promo_badge': promo_badge}

    def scrape_coles_page(self, page, url):
        page.goto(url, wait_until='domcontentloaded', timeout=60000)
        
        if self.debug_var.get():
            self._save_debug_html(page, 'coles', url)

        # Check for CAPTCHA and wait for user to solve it
        try:
            captcha_locator = page.locator('iframe[title="Widget containing a Cloudflare security challenge"]')
            captcha_locator.wait_for(timeout=5000) # Quick check to see if it's there
            
            # If the above line doesn't throw an error, the CAPTCHA is present
            self.log("!!! ACTION REQUIRED: CAPTCHA detected. Please solve the challenge in the browser window.")
            self.log("    The script will wait up to 2 minutes for you to complete it...")
            
            # Wait for the product page to load after CAPTCHA
            page.wait_for_selector('h1[data-testid="title"], section[data-testid="product_price"]', timeout=120000)
            self.log("   CAPTCHA solved! Resuming scraping.")

        except PlaywrightTimeoutError:
            # This is the normal path - the CAPTCHA was NOT found, so we proceed.
            self.log("   No CAPTCHA detected, proceeding with scrape.")
            # Wait for the product title or price section to appear
            page.wait_for_selector('h1[data-testid="title"], section[data-testid="product_price"]', timeout=20000)
        
        name = "Not found"
        try:
            # Get product name from the h1 title
            name = page.locator('h1[data-testid="title"]').inner_text(timeout=5000)
        except Exception:
            pass
        
        price, was_price, cup_price, promo_badge = "Not found", "Not applicable", "Not found", ""
        
        try: 
            # Get price from the pricing span
            price = page.locator('span[data-testid="pricing"]').inner_text(timeout=5000).replace('$', '').strip()
        except Exception: 
            pass
        
        try: 
            was_price_full = page.locator('.price__was').inner_text(timeout=2000)
            was_price = re.search(r'[\d.]+', was_price_full).group()
        except Exception: 
            pass
        
        try: 
            # Get unit price from the calculation method div
            cup_price = page.locator('.price__calculation_method').inner_text(timeout=2000).strip()
        except Exception: 
            pass
        
        if was_price != "Not applicable":
            promo_badge = "Special" 
            if was_price and price:
                try:
                    if float(was_price) / 2 == float(price): 
                        promo_badge = "1/2 Price"
                except Exception: 
                    pass
        
        return {'store': 'Coles', 'name': name, 'price': price, 'was_price': was_price, 'cup_price': cup_price, 'url': url, 'promo_badge': promo_badge}

    def scraping_thread(self):
        all_urls = list(self.url_listbox.get(0, tk.END))
        urls_to_scrape = []
        scrape_ww = self.scrape_woolworths.get()
        scrape_cl = self.scrape_coles.get()
        for url in all_urls:
            if 'woolworths.com.au' in url and scrape_ww: urls_to_scrape.append(url)
            elif 'coles.com.au' in url and scrape_cl: urls_to_scrape.append(url)
        if not urls_to_scrape:
            self.log("No URLs to scrape based on current selection!")
            self.is_scraping = False
            self.scrape_button.config(text="Start Scraping", state='normal')
            return
            
        self.scraped_data, self.progress['value'] = [], 0
        self.tree.delete(*self.tree.get_children())
        self.progress['maximum'] = len(urls_to_scrape)
        self.log(f"Starting scraper for {len(urls_to_scrape)} URLs...")

        # Enhanced stealth script
        stealth_script = """
        // Override webdriver property
        Object.defineProperty(navigator, 'webdriver', {
            get: () => undefined
        });
        
        // Override plugins to look more realistic
        Object.defineProperty(navigator, 'plugins', {
            get: () => [1, 2, 3, 4, 5]
        });
        
        // Override language properties
        Object.defineProperty(navigator, 'languages', {
            get: () => ['en-AU', 'en']
        });
        
        // Fix chrome runtime
        window.chrome = {
            runtime: {}
        };
        
        // Override permissions query
        const originalQuery = window.navigator.permissions.query;
        window.navigator.permissions.query = (parameters) => (
            parameters.name === 'notifications' ?
                Promise.resolve({ state: Notification.permission }) :
                originalQuery(parameters)
        );
        """

        try:
            with sync_playwright() as p:
                # Launch with more realistic browser arguments
                browser_args = []
                if not self.headless_var.get():
                    browser_args = [
                        '--disable-blink-features=AutomationControlled',
                        '--disable-dev-shm-usage',
                        '--no-sandbox',
                        '--disable-web-security',
                        '--disable-features=IsolateOrigins,site-per-process'
                    ]
                
                browser = p.chromium.launch(
                    headless=self.headless_var.get(),
                    args=browser_args
                )
                
                # More complete context setup
                context = browser.new_context(
                    user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
                    viewport={'width': 1920, 'height': 1080},
                    screen={'width': 1920, 'height': 1080},
                    locale='en-AU',
                    timezone_id='Australia/Brisbane',
                    geolocation={'latitude': -27.4698, 'longitude': 153.0251},
                    permissions=['geolocation'],
                    device_scale_factor=1,
                    has_touch=False,
                    is_mobile=False
                )
                
                context.add_init_script(stealth_script)
                context.grant_permissions(['geolocation'], origin='https://www.coles.com.au')
                context.grant_permissions(['geolocation'], origin='https://www.woolworths.com.au')
                
                # Load cookies if available
                self.load_cookies(context)
                
                page = context.new_page()
                
                # Warmup browser for Coles if we're scraping Coles URLs
                has_coles = any('coles.com.au' in url for url in urls_to_scrape)
                if has_coles and not self.headless_var.get():
                    self.warmup_browser(page)

                for i, url in enumerate(urls_to_scrape, 1):
                    self.log(f"Scraping {i}/{len(urls_to_scrape)}: {url.split('/')[-1]}")
                    data = {}
                    try:
                        if 'woolworths.com.au' in url: 
                            data = self.scrape_woolworths_page(page, url)
                        elif 'coles.com.au' in url: 
                            data = self.scrape_coles_page(page, url)
                        else: 
                            data = {'error': 'Unknown store', 'url': url}
                    except Exception as e:
                        data = {'error': str(e), 'url': url}
                        
                    self.scraped_data.append(data)
                    
                    if 'error' not in data:
                        _, promo_type = self.calculate_discount(data['price'], data['was_price'], data.get('promo_badge', ''))
                        price_display = f"${data['price']}" if data['price'] != "Not found" else "N/A"
                        was_display = f"${data['was_price']}" if data['was_price'] != "Not applicable" else "-"
                        self.tree.insert('', tk.END, values=(data['store'], data['name'], price_display, was_display, data['cup_price'], promo_type or ""), tags=(data['url'],))
                        self.log(f"  ✓ {data['store']}: {data['name']}")
                    else:
                        self.log(f"  ✗ Error for {url}: {data.get('error', 'Unknown error')}")
                        
                    self.progress['value'] = i
                    self.root.update_idletasks()
                    
                    if i < len(urls_to_scrape):
                        # Longer, more variable delays for Coles
                        next_url = urls_to_scrape[i] if i < len(urls_to_scrape) else ""
                        if 'coles.com.au' in next_url:
                            sleep_time = random.uniform(5, 10)  # Longer for Coles
                        else:
                            sleep_time = random.uniform(2, 5)   # Original for Woolworths
                        
                        self.log(f"  Waiting for {sleep_time:.1f} seconds before next product...")
                        time.sleep(sleep_time)

                # Save cookies before closing
                self.save_cookies(context)
                browser.close()
                
        except Exception as e:
            self.log(f"An unexpected error occurred during scraping: {str(e)}")

        successful = sum(1 for d in self.scraped_data if 'error' not in d)
        self.summary_label.config(text=f"Scraped: {successful}/{len(urls_to_scrape)} products")
        if successful > 0:
            self.csv_button.config(state='normal')
            if HAS_EXCEL: self.excel_button.config(state='normal')
            if HAS_GEMINI: self.ai_button.config(state='normal')
        self.log("Scraping complete!")
        self.is_scraping = False
        self.scrape_button.config(text="Start Scraping", state='normal')
        
    def on_item_double_click(self, event):
        item_id = self.tree.identify_row(event.y)
        if item_id:
            tags = self.tree.item(item_id, "tags")
            if tags and tags[0]:
                webbrowser.open(tags[0])
                self.log(f"Opened URL in browser: {tags[0]}")

    def start_ai_analysis(self):
        if not HAS_GEMINI:
            messagebox.showerror("Gemini Not Available", "The 'google-generativeai' library is not installed.\nPlease run: pip install google-generativeai")
            return
        if not self.api_key.get():
            messagebox.showerror("API Key Missing", "Please go to the Settings tab and enter your Gemini API key.")
            return
        ai_window = tk.Toplevel(self.root)
        ai_window.title(f"Gemini AI Analysis ({self.model_var.get()})")
        ai_window.geometry("800x600")
        results_text = scrolledtext.ScrolledText(ai_window, height=20, width=80, wrap=tk.WORD, font=("Arial", 10))
        results_text.pack(padx=10, pady=10, fill='both', expand=True)
        results_text.insert('1.0', "Preparing data and contacting Gemini API... Please wait.")
        threading.Thread(target=self.run_gemini_analysis_thread, args=(results_text,), daemon=True).start()

    def run_gemini_analysis_thread(self, results_text_widget):
        model_name = ""
        try:
            genai.configure(api_key=self.api_key.get())
            model_name = self.model_var.get()
            model = genai.GenerativeModel(model_name)
            data_str = "Product Pricing Data from Woolworths and Coles:\n" + "-"*50 + "\n"
            for item in self.scraped_data:
                if 'error' not in item:
                    _, promo_type = self.calculate_discount(item['price'], item['was_price'], item.get('promo_badge', ''))
                    data_str += f"\nStore: {item['store']}\n"
                    data_str += f"  Product: {item['name']}\n"
                    data_str += f"  Current Price: ${item['price']}\n"
                    if item['was_price'] != "Not applicable":
                        data_str += f"  Was Price: ${item['was_price']}\n"
                        data_str += f"  Promotion: {promo_type}\n"
                    data_str += f"  Unit Price: {item['cup_price']}\n"
            prompt = f"""You are a market analyst for a consumer goods company. Analyze the following competitive pricing data for detergent sheets from both Woolworths and Coles in Australia.
**Data:**
{data_str}
**Your Task:**
Provide a concise but insightful analysis covering these points:
1. **Cross-Store Price Comparison:** Are there significant price differences for similar products between Woolworths and Coles? Which store appears to be cheaper overall for this product category?
2. **Best Value:** Based on unit price (e.g., price per sheet/load), which specific products offer the best value for money, regardless of the store?
3. **Promotional Strategy:** Compare the promotional activities between the two stores. Is one more aggressive with discounts?
4. **Market Opportunities:** Identify any gaps in the market. For example, are there product sizes or types available at one store but not the other?
5. **Strategic Recommendations:** Provide one key recommendation for a brand selling in both stores. How should they tailor their pricing or promotional strategy for each retailer?
Structure your response with clear headings. Be professional and data-driven."""
            if self.debug_var.get():
                self.log("--- DEBUG: AI PROMPT ---\n" + prompt + "\n--- END AI PROMPT ---")
            response = model.generate_content(prompt)
            results_text_widget.delete('1.0', tk.END)
            results_text_widget.insert('1.0', response.text)
        except Exception as e:
            error_message = f"An error occurred during AI analysis:\n\n{str(e)}"
            results_text_widget.delete('1.0', tk.END)
            results_text_widget.insert('1.0', error_message)
            if self.debug_var.get():
                self.log(f"--- DEBUG: AI ANALYSIS FAILED ---\nModel used: {model_name}\nError Type: {type(e).__name__}\nFull Traceback:\n{traceback.format_exc()}--- END DEBUG ---")

    def save_settings(self):
        key, model = self.api_key.get().strip(), self.model_var.get()
        if not key:
            messagebox.showwarning("Empty Key", "API key field is empty.")
            return
        try:
            with open("config.json", "w") as f:
                json.dump({"api_key": key, "model_name": model}, f, indent=4)
            self.log("Settings saved successfully.")
            messagebox.showinfo("Success", "Settings saved successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Could not save settings: {e}")

    def load_settings(self):
        try:
            if os.path.exists("config.json"):
                with open("config.json", "r") as f:
                    settings = json.load(f)
                    self.api_key.set(settings.get("api_key", ""))
                    saved_model = settings.get("model_name", self.available_models[0])
                    self.model_var.set(saved_model if saved_model in self.available_models else self.available_models[0])
                    if self.api_key.get(): self.log("Loaded settings from config.json")
        except Exception as e:
            self.log(f"Could not load settings: {e}")
            self.api_key.set("")
            self.model_var.set(self.available_models[0])

    def start_scraping(self):
        if self.is_scraping: return
        self.is_scraping = True
        for button in [self.scrape_button, self.csv_button, self.excel_button, self.ai_button]:
            button.config(state='disabled')
        self.scrape_button.config(text="Scraping...")
        threading.Thread(target=self.scraping_thread, daemon=True).start()

    def export_csv(self):
        if not self.scraped_data: return
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            initialfile=f"price_comparison_{timestamp}.csv"
        )
        if not filename: return
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                fieldnames = ['Store', 'Product Name', 'Current Price', 'Was Price', 'Unit Price', 'Promotion', 'URL']
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                for item in (d for d in self.scraped_data if 'error' not in d):
                    _, promo_type = self.calculate_discount(item['price'], item['was_price'], item.get('promo_badge', ''))
                    writer.writerow({
                        'Store': item['store'], 'Product Name': item['name'], 'Current Price': item['price'], 
                        'Was Price': item['was_price'], 'Unit Price': item['cup_price'], 
                        'Promotion': promo_type, 'URL': item['url']
                    })
            self.log(f"CSV exported to {filename}")
            messagebox.showinfo("Success", f"Data exported to {filename}")
        except Exception as e: 
            messagebox.showerror("Error", f"Failed to export CSV: {e}")

    def export_excel(self):
        if not HAS_EXCEL:
            messagebox.showwarning("Excel Not Available", "Please install openpyxl:\npip install openpyxl")
            return
        if not self.scraped_data: return
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"price_comparison_{timestamp}.xlsx"
        )
        if not filename: return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Price Comparison"
            ws.append(['Store', 'Product Name', 'Current Price', 'Was Price', 'Unit Price', 'Promotion', 'URL'])
            header_font, header_fill = Font(bold=True, color="FFFFFF"), PatternFill(start_color="366092", fill_type="solid")
            for cell in ws[1]: cell.font, cell.fill = header_font, header_fill
            half_price_fill, special_fill = PatternFill(start_color="FFC7CE", fill_type="solid"), PatternFill(start_color="FFEB9C", fill_type="solid")
            for item in (d for d in self.scraped_data if 'error' not in d):
                _, promo_type = self.calculate_discount(item['price'], item['was_price'], item.get('promo_badge', ''))
                try: price = float(item['price']) 
                except (ValueError, TypeError): price = ""
                try: was_price = float(item['was_price'])
                except (ValueError, TypeError): was_price = ""
                ws.append([item['store'], item['name'], price, was_price, item['cup_price'], promo_type, item['url']])
                if "HALF PRICE" in str(promo_type).upper():
                    for cell in ws[ws.max_row]: cell.fill = half_price_fill
                elif promo_type:
                    for cell in ws[ws.max_row]: cell.fill = special_fill
            for column in ws.columns:
                max_length = max((len(str(cell.value)) for cell in column if cell.value is not None), default=0)
                ws.column_dimensions[column[0].column_letter].width = min((max_length + 2), 60)
            wb.save(filename)
            self.log(f"Excel file exported to {filename}")
            messagebox.showinfo("Success", f"Data exported to {filename}")
        except Exception as e: messagebox.showerror("Error", f"Failed to export Excel: {e}")

def main():
    root = tk.Tk()
    app = MultiStoreScraperGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()
